import csv
import sys
from plistlib import InvalidFileException

from openpyxl import load_workbook


def main():
    # Check if the correct number of arguments are provided
    if len(sys.argv) != 2:
        print(
            "Usage: python visit_type_excel_converter.py <./string/excel/file_dir.xlsx>"
        )
        sys.exit(1)

    excel_file = sys.argv[1]
    print(f"Received dir parameter: '{excel_file}'")

    try:
        workbook = load_workbook(filename=excel_file)
        sheet = workbook.active
        print("Found the Excel file to convert :)")
    except InvalidFileException as e:
        print(f"Error loading Excel file: {e}")
        sys.exit(1)
    except PermissionError:
        print(f"Error: Permission denied to open the file '{excel_file}'.")
        sys.exit(1)

    excel_data = []
    for row in sheet.iter_rows(
        min_row=2, values_only=True
    ):  # Assuming the first row is header
        excel_data.append(row)

    # Filter out rows with None in the second column
    excel_data = [row for row in excel_data if row[1] is not None]

    if len(excel_data) > 0:
        print("Filtered Excel data loaded successfully")
    else:
        print("No valid data found in the Excel file")

    # Load the CSV file for the example structure
    try:
        with open(
            "./datafiles/sponsor_library/visit/visit_type_exp.csv",
            mode="r",
            encoding="utf-8",
        ) as csv_file:
            csv_reader = csv.reader(csv_file)
            csv_data = list(csv_reader)
        print("Found the CSV with example structure :)")
    except FileNotFoundError:
        print("Error: The CSV file 'visit_type_exp.csv' does not exist.")
        sys.exit(1)
    except PermissionError:
        print("Error: Permission denied to read the CSV file.")
        sys.exit(1)

    if len(csv_data) > 0:
        print("CSV data loaded successfully")
    else:
        print("No data found in the CSV file")

    new_csv = []
    for row in excel_data:
        sponsor_name = row[1]
        epoch_type = row[-1]
        new_csv.append(
            {
                "CT_CD_LIST_SUBMVAL": sponsor_name,
                "CT_NAME": sponsor_name,
                "NAME_SENTENSE_CASE": sponsor_name,
                "CT_SUBMVAL": sponsor_name,
                "DEFINITION": sponsor_name + " visit",
                "ORDER": 1,
                "EPOCHS": epoch_type,
            }
        )

    # Write the result to a CSV file
    output_csv_file = "./output_result.csv"
    try:
        with open(output_csv_file, mode="w", newline="", encoding="utf-8") as csvfile:
            fieldnames = new_csv[0]  # header
            writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
            writer.writeheader()
            writer.writerows(new_csv)

        print(f"Result has been written to '{output_csv_file}'")
    except PermissionError:
        print(f"Error: Permission denied to write to the file '{output_csv_file}'.")
        sys.exit(1)
    except IOError:
        print(
            f"Error: An I/O error occurred while writing to the file '{output_csv_file}'."
        )
        sys.exit(1)

    for item in new_csv:
        print(item)


if __name__ == "__main__":
    main()
